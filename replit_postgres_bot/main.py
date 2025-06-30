import logging
import os
import datetime
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
from telegram.constants import ChatAction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import matplotlib.pyplot as plt
from dotenv import load_dotenv

load_dotenv()

# –õ–æ–≥–∏—Ä–æ–≤–∞–Ω–∏–µ
logger = logging.getLogger(__name__)
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# –ö–æ–Ω—Å—Ç–∞–Ω—Ç—ã –∏–∑ .env
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", "8139881064"))
CHANNEL_ID = os.getenv("CHANNEL_ID", "@F_S_Ta")

CUR, NEW, COST, PERIOD = range(4)
HISTORY_FILE = "–∏—Å—Ç–æ—Ä–∏—è.xlsx"

async def check_subscription(update: Update) -> bool:
    user_id = update.effective_user.id
    try:
        member = await update.get_bot().get_chat_member(CHANNEL_ID, user_id)
        return member.status in ["member", "administrator", "creator"]
    except:
        return False

async def send_subscription_prompt(update: Update) -> None:
    keyboard = [
        [InlineKeyboardButton("üì∫ –ü–æ–¥–ø–∏—Å–∞—Ç—å—Å—è –Ω–∞ –∫–∞–Ω–∞–ª", url=f"https://t.me/{CHANNEL_ID.lstrip('@')}")],
        [InlineKeyboardButton("‚úÖ –Ø –ø–æ–¥–ø–∏—Å–∞–ª—Å—è, –ø—Ä–æ–≤–µ—Ä–∏—Ç—å", callback_data="check_subscription")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        "üîí –î–ª—è –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è –±–æ—Ç–∞ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –ø–æ–¥–ø–∏—Å–∞—Ç—å—Å—è –Ω–∞ –Ω–∞—à –∫–∞–Ω–∞–ª!\n\n"
        "–ü–æ—Å–ª–µ –ø–æ–¥–ø–∏—Å–∫–∏ –Ω–∞–∂–º–∏—Ç–µ –∫–Ω–æ–ø–∫—É '–Ø –ø–æ–¥–ø–∏—Å–∞–ª—Å—è':",
        reply_markup=reply_markup
    )

async def subscription_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    
    if query.data == "check_subscription":
        if await check_subscription(update):
            await query.edit_message_text("‚úÖ –û—Ç–ª–∏—á–Ω–æ! –¢–µ–ø–µ—Ä—å –≤—ã –º–æ–∂–µ—Ç–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç—å—Å—è –±–æ—Ç–æ–º.\n\n–í–≤–µ–¥–∏—Ç–µ —Ç–µ–∫—É—â–∏–π —Ç–∞—Ä–∏—Ñ (‚ÇΩ/–º–µ—Å):")
            return CUR
        else:
            keyboard = [
                [InlineKeyboardButton("üì∫ –ü–æ–¥–ø–∏—Å–∞—Ç—å—Å—è –Ω–∞ –∫–∞–Ω–∞–ª", url=f"https://t.me/{CHANNEL_ID.lstrip('@')}")],
                [InlineKeyboardButton("‚úÖ –Ø –ø–æ–¥–ø–∏—Å–∞–ª—Å—è, –ø—Ä–æ–≤–µ—Ä–∏—Ç—å", callback_data="check_subscription")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.edit_message_text(
                "‚ùå –í—ã –µ—â—ë –Ω–µ –ø–æ–¥–ø–∏—Å–∞–Ω—ã –Ω–∞ –∫–∞–Ω–∞–ª.\n\n"
                "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–¥–ø–∏—à–∏—Ç–µ—Å—å –∏ –Ω–∞–∂–º–∏—Ç–µ –∫–Ω–æ–ø–∫—É –ø–æ–≤—Ç–æ—Ä–Ω–æ:",
                reply_markup=reply_markup
            )
            return ConversationHandler.END
    
    return ConversationHandler.END

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await check_subscription(update):
        await send_subscription_prompt(update)
        return ConversationHandler.END

    await update.message.reply_text("–í–≤–µ–¥–∏—Ç–µ —Ç–µ–∫—É—â–∏–π —Ç–∞—Ä–∏—Ñ (‚ÇΩ/–º–µ—Å):")
    return CUR

async def cur_tariff(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['cur'] = float(update.message.text)
    except ValueError:
        await update.message.reply_text("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏—Ç–µ —á–∏—Å–ª–æ.")
        return CUR
    await update.message.reply_text("–í–≤–µ–¥–∏—Ç–µ –∞/–ø–ª–∞—Ç—É –¥–ª—è –Ω–æ–≤–æ–≥–æ —Ç–∞—Ä–∏—Ñ–∞ (‚ÇΩ/–º–µ—Å):")
    return NEW

async def new_tariff(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['new'] = float(update.message.text)
    except ValueError:
        await update.message.reply_text("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏—Ç–µ —á–∏—Å–ª–æ.")
        return NEW
    await update.message.reply_text("–í–≤–µ–¥–∏—Ç–µ —Å—Ç–æ–∏–º–æ—Å—Ç—å –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è  (—Ä—É–±):")
    return COST

async def cost(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['cost'] = float(update.message.text)
    except ValueError:
        await update.message.reply_text("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏—Ç–µ —á–∏—Å–ª–æ.")
        return COST
    await update.message.reply_text("–ó–∞ –∫–∞–∫–æ–π –ø–µ—Ä–∏–æ–¥ —Ä–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º –≤—ã–≥–æ–¥—É (–ª–µ—Ç)?")
    return PERIOD

def format_table_for_telegram(rows, summary):
    """–§–æ—Ä–º–∞—Ç–∏—Ä—É–µ—Ç —Ç–∞–±–ª–∏—Ü—É –¥–ª—è –∫—Ä–∞—Å–∏–≤–æ–≥–æ –æ—Ç–æ–±—Ä–∞–∂–µ–Ω–∏—è –≤ Telegram"""
    # –°–æ–∑–¥–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–æ–∫ —Ç–∞–±–ª–∏—Ü—ã
    table_text = "üìä **–†–ê–°–ß–ï–¢ –¢–ê–†–ò–§–û–í**\n\n"
    table_text += f"```\n"
    table_text += f"{'–ú–µ—Å':<4} {'–°—Ç–∞—Ä–∞—è':<8} {'–ù–æ–≤–∞—è':<8} {'–≠–∫–æ–Ω–æ–º–∏—è':<8}\n"
    table_text += f"{'-'*4} {'-'*8} {'-'*8} {'-'*8}\n"
    
    # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç—Ä–æ–∫–∏ –¥–∞–Ω–Ω—ã—Ö (–ø–æ–∫–∞–∑—ã–≤–∞–µ–º —Ç–æ–ª—å–∫–æ –ø–µ—Ä–≤—ã–µ 12 –∏ –ø–æ—Å–ª–µ–¥–Ω–∏–µ 12 –º–µ—Å—è—Ü–µ–≤ –µ—Å–ª–∏ –±–æ–ª—å—à–µ 24)
    data_rows = rows[1:]  # —É–±–∏—Ä–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–æ–∫
    
    if len(data_rows) <= 24:
        for row in data_rows:
            month, old, new, diff = row
            diff_sign = "+" if diff >= 0 else ""
            table_text += f"{month:<4} {old:<8} {new:<8} {diff_sign}{diff:<7}\n"
    else:
        # –ü–µ—Ä–≤—ã–µ 12 –º–µ—Å—è—Ü–µ–≤
        for row in data_rows[:12]:
            month, old, new, diff = row
            diff_sign = "+" if diff >= 0 else ""
            table_text += f"{month:<4} {old:<8} {new:<8} {diff_sign}{diff:<7}\n"
        
        table_text += f"{'...':<4} {'...':<8} {'...':<8} {'...':<8}\n"
        
        # –ü–æ—Å–ª–µ–¥–Ω–∏–µ 12 –º–µ—Å—è—Ü–µ–≤
        for row in data_rows[-12:]:
            month, old, new, diff = row
            diff_sign = "+" if diff >= 0 else ""
            table_text += f"{month:<4} {old:<8} {new:<8} {diff_sign}{diff:<7}\n"
    
    table_text += f"```\n\n"
    table_text += f"üìà **–ò–¢–û–ì–ò:**\n{summary}"
    
    return table_text

async def generate_and_send_reports(cur, new, cost, period_years, user_name, bot, chat_id, admin_id=None):
    months = int(period_years * 12)
    cumulative_old = 0
    cumulative_new = cost
    payback_month = None
    rows = [["–ú–µ—Å—è—Ü", "–°—Ç–∞—Ä–∞—è", "–ù–æ–≤–∞—è", "–≠–∫–æ–Ω–æ–º–∏—è"]]

    for m in range(1, months + 1):
        cumulative_old += cur
        cumulative_new += new
        diff = cumulative_old - cumulative_new
        rows.append([m, round(cumulative_old), round(cumulative_new), round(diff)])
        if payback_month is None and diff >= 0:
            payback_month = m

    summary = f"–û–∫—É–ø–∞–µ–º–æ—Å—Ç—å: {payback_month} –º–µ—Å.\n" if payback_month else "–û–∫—É–ø–∞–µ–º–æ—Å—Ç—å –Ω–µ –¥–æ—Å—Ç–∏–≥–Ω—É—Ç–∞.\n"
    summary += f"–û–±—â–∞—è —ç–∫–æ–Ω–æ–º–∏—è –∑–∞ {months} –º–µ—Å.: {round(cumulative_old - cumulative_new)}‚ÇΩ"

    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Ç–∞–±–ª–∏—Ü—É –≤ Telegram
    table_message = format_table_for_telegram(rows, summary)
    await bot.send_message(chat_id=chat_id, text=table_message, parse_mode='Markdown')

    # PDF
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    style = getSampleStyleSheet()["BodyText"]
    pdf_table = Table(rows)
    styles = [('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
              ('GRID', (0, 0), (-1, -1), 1, colors.grey),
              ('ALIGN', (0, 0), (-1, -1), 'CENTER')]
    for idx, row in enumerate(rows[1:], start=1):
        if row[3] < 0:
            styles.append(('TEXTCOLOR', (0, idx), (-1, idx), colors.red))
    pdf_table.setStyle(styles)
    intro_text = Paragraph(
        f"<b>–ò—Å—Ö–æ–¥–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ:</b><br/>"
        f"–¢–µ–∫—É—â–∏–π —Ç–∞—Ä–∏—Ñ: {cur:.2f}‚ÇΩ/–º–µ—Å<br/>"
        f"–ù–æ–≤—ã–π —Ç–∞—Ä–∏—Ñ: {new:.2f}‚ÇΩ/–º–µ—Å<br/>"
        f"–ü–æ–¥–∫–ª—é—á–µ–Ω–∏–µ: {cost:.2f}‚ÇΩ<br/>"
        f"–ü–µ—Ä–∏–æ–¥: {months} –º–µ—Å.<br/><br/><b>–ò—Ç–æ–≥–∏:</b><br/>{summary.replace(chr(10), '<br/>')}",
        style)
    doc.build([intro_text, pdf_table])
    pdf_buffer.seek(0)
    await bot.send_document(chat_id=chat_id, document=pdf_buffer, filename="–≤—ã–≥–æ–¥–∞.pdf")

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "–†–∞—Å—á—ë—Ç —Ç–∞—Ä–∏—Ñ–∞"
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    negative_fill = PatternFill("solid", fgColor="FFC7CE")
    ws.append(["–ò—Å—Ö–æ–¥–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ"])
    ws.append(["–¢–µ–∫—É—â–∏–π —Ç–∞—Ä–∏—Ñ", cur])
    ws.append(["–ù–æ–≤—ã–π —Ç–∞—Ä–∏—Ñ", new])
    ws.append(["–°—Ç–æ–∏–º–æ—Å—Ç—å –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è", cost])
    ws.append(["–ü–µ—Ä–∏–æ–¥ (–º–µ—Å)", months])
    ws.append([])
    ws.append(rows[0])
    for col in range(1, 5):
        cell = ws.cell(row=7, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
    for idx, row in enumerate(rows[1:], start=1):
        ws.append(row)
        if row[3] < 0:
            for col in range(1, 5):
                ws.cell(row=7 + idx, column=col).fill = negative_fill
    chart = LineChart()
    chart.title = "–ì—Ä–∞—Ñ–∏–∫ –Ω–∞–∫–æ–ø–ª–µ–Ω–Ω—ã—Ö –∑–∞—Ç—Ä–∞—Ç"
    chart.y_axis.title = '‚ÇΩ'
    chart.x_axis.title = '–ú–µ—Å—è—Ü'
    data = Reference(ws, min_col=2, min_row=7, max_col=3, max_row=7 + months)
    cats = Reference(ws, min_col=1, min_row=8, max_row=7 + months)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F8")
    plt.figure(figsize=(1, 1))
    plt.plot([0], [0], 'o', markersize=30, color='green')
    plt.axis('off')
    img_buf = BytesIO()
    plt.savefig(img_buf, format='png', bbox_inches='tight', transparent=True)
    img_buf.seek(0)
    logo_img = Image(img_buf)
    logo_img.width = 60
    logo_img.height = 60
    ws.add_image(logo_img, "F1")
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    await bot.send_document(chat_id=chat_id, document=excel_buffer, filename="–≤—ã–≥–æ–¥–∞.xlsx")

    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –∏—Å—Ç–æ—Ä–∏—é
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    if not os.path.exists(HISTORY_FILE):
        wb_hist = Workbook()
        ws_hist = wb_hist.active
        ws_hist.title = "–ò—Å—Ç–æ—Ä–∏—è"
        ws_hist.append(["–î–∞—Ç–∞", "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å", "–¢–µ–∫—É—â–∏–π —Ç–∞—Ä–∏—Ñ", "–ù–æ–≤—ã–π —Ç–∞—Ä–∏—Ñ", "–ü–æ–¥–∫–ª—é—á–µ–Ω–∏–µ", "–ü–µ—Ä–∏–æ–¥", "–û–∫—É–ø–∞–µ–º–æ—Å—Ç—å (–º–µ—Å)", "–≠–∫–æ–Ω–æ–º–∏—è"])
    else:
        wb_hist = load_workbook(HISTORY_FILE)
        ws_hist = wb_hist.active
    ws_hist.append([
        now,
        user_name,
        cur, new, cost, months,
        payback_month if payback_month else "–ù–µ –¥–æ—Å—Ç–∏–≥–Ω—É—Ç–∞",
        round(cumulative_old - cumulative_new)
    ])
    wb_hist.save(HISTORY_FILE)

    # –û—Ç–ø—Ä–∞–≤–∫–∞ –∞–¥–º–∏–Ω—É
    if admin_id:
        # PDF –∏ Excel —É–∂–µ –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é, –ø–æ—ç—Ç–æ–º—É —Å–±—Ä–∞—Å—ã–≤–∞–µ–º –∫—É—Ä—Å–æ—Ä—ã
        pdf_buffer.seek(0)
        excel_buffer.seek(0)
        await bot.send_document(chat_id=admin_id, document=pdf_buffer, filename="user_–≤—ã–≥–æ–¥–∞.pdf")
        await bot.send_document(chat_id=admin_id, document=excel_buffer, filename="user_–≤—ã–≥–æ–¥–∞.xlsx")

async def period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['period'] = float(update.message.text)
    except ValueError:
        await update.message.reply_text("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏—Ç–µ —á–∏—Å–ª–æ.")
        return PERIOD

    await generate_and_send_reports(
        cur=context.user_data['cur'],
        new=context.user_data['new'],
        cost=context.user_data['cost'],
        period_years=context.user_data['period'],
        user_name=update.effective_user.full_name,
        bot=context.bot,
        chat_id=update.effective_chat.id,
        admin_id=ADMIN_ID
    )
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text('–†–∞—Å—á—ë—Ç –æ—Ç–º–µ–Ω—ë–Ω.')
    return ConversationHandler.END

async def history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await check_subscription(update):
        await send_subscription_prompt(update)
        return
        
    user_name = update.effective_user.full_name
    if not os.path.exists(HISTORY_FILE):
        await update.message.reply_text("–ò—Å—Ç–æ—Ä–∏—è –ø—É—Å—Ç–∞.")
        return
    wb_src = load_workbook(HISTORY_FILE)
    ws_src = wb_src.active
    wb_user = Workbook()
    ws_user = wb_user.active
    ws_user.title = "–ò—Å—Ç–æ—Ä–∏—è"
    ws_user.append([cell.value for cell in ws_src[1]])
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if row[1] == user_name:
            ws_user.append(row)
    if ws_user.max_row == 1:
        await update.message.reply_text("–£ –≤–∞—Å –ø–æ–∫–∞ –Ω–µ—Ç —Ä–∞—Å—á—ë—Ç–æ–≤.")
        return
    buf = BytesIO()
    wb_user.save(buf)
    buf.seek(0)
    await update.message.reply_document(buf, filename="–º–æ—è_–∏—Å—Ç–æ—Ä–∏—è.xlsx")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    help_text = (
        "–î–æ—Å—Ç—É–ø–Ω—ã–µ –∫–æ–º–∞–Ω–¥—ã:\n"
        "/start - –ù–∞—á–∞—Ç—å —Ä–∞—Å—á–µ—Ç —Ç–∞—Ä–∏—Ñ–∞\n"
        "/cancel - –û—Ç–º–µ–Ω–∏—Ç—å —Ç–µ–∫—É—â–∏–π –≤–≤–æ–¥\n"
        "/history - –ü–æ–∫–∞–∑–∞—Ç—å –≤–∞—à—É –∏—Å—Ç–æ—Ä–∏—é —Ä–∞—Å—á–µ—Ç–æ–≤\n"
        "/help - –ü–æ–∫–∞–∑–∞—Ç—å —ç—Ç–æ —Å–æ–æ–±—â–µ–Ω–∏–µ"
    )
    await update.message.reply_text(help_text)

  # if __name__ == '__main__':
   
    # –ó–∞–ø—É—Å–∫ —Å–µ—Ä–≤–µ—Ä–∞ –≤ –æ—Ç–¥–µ–ª—å–Ω–æ–º –ø–æ—Ç–æ–∫–µ

import asyncio
from http_server import run_http_server
from telegram.ext import ApplicationBuilder, CommandHandler
import os
from dotenv import load_dotenv

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")

async def start(update, context):
    await update.message.reply_text("–ü—Ä–∏–≤–µ—Ç! –ë–æ—Ç —Ä–∞–±–æ—Ç–∞–µ—Ç.")

async def main():
    asyncio.create_task(run_http_server())

    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))

    await app.run_polling()

if __name__ == '__main__':
    asyncio.run(main())    
    # –ó–∞–ø—É—Å–∫ Telegram –±–æ—Ç–∞
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    from telegram.ext import CallbackQueryHandler
    
    conv = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            CUR: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, cur_tariff),
                CallbackQueryHandler(subscription_callback, pattern="check_subscription")
            ],
            NEW: [MessageHandler(filters.TEXT & ~filters.COMMAND, new_tariff)],
            COST: [MessageHandler(filters.TEXT & ~filters.COMMAND, cost)],
            PERIOD: [MessageHandler(filters.TEXT & ~filters.COMMAND, period)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    app.add_handler(conv)
    app.add_handler(CommandHandler("history", history))
    app.add_handler(CommandHandler("help", help_command))
    app.run_polling()
        
